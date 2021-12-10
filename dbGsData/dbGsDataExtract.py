#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2021/9/26 11:14
# @Author  : aiyingyi
# @FileName: dbGsDataExtract.py
# @Software: PyCharm

'''

达标公示数据爬取，注意不是达标公布数据，公示数据没有达标编号，详情页连接是从excel中读取的

将爬取数据的连接直接放在excel的第一列

'''

import openpyxl
import requests

from bs4 import BeautifulSoup
from utils.spiderUtils import utils


# 返回一个综合数据的模板对象
def get_data():
    return {
        'CPXH': '', 'SCQY': '', 'CPMC': '', 'CLLB': '', 'RLZL': '', 'DBCXBH': '',
        'PC': '', 'LEVEL': '', 'KCLX': '', 'DPXH': '', 'OUT_SIZE': '', 'ZZL': '', 'QDXS': '',
        'LTGG': '', 'CMFHZZ': '', 'QXBFHZZ': '', 'JGDSL_QX': '', 'JGDSL_SPCZM': '', 'ZHBZBS': '',
        'ZDQXS': '', 'ZDJXZDTZZZ': '', 'WDJKZZ': '', 'JSSLTBS': '', 'LGBZXH': '', 'ZDQCP': '',
        'FBZDZZ': '', 'ESCXH': '', 'LDWSXH': '', 'ZHRLXHL': '', 'WXHWCLLX': '', 'PQGCKWZ': '', 'HSQZDZZ': '',
        'DWLHSQBJXT': '', 'QFBHZZ': '', 'GTWBZHBF': '', 'TDYSJZ': '', 'FDJXH': '', 'HXLBNCC_RJ': '',
        'ZBZL': '', 'ZGCS': '', 'ZXZ': '', 'LTSL': '', 'HXBFH': '', 'QZWBJSBZ': '',
        'QCDJDXJ': '', 'RQPSLJWZ': '', 'QYZDXT': '', 'YLCSLJQ_CQT': '', 'YLCSLJQ_ZDQS': '', 'WXDW': '',
        'QTXLBJ': '', 'LTQYJCXTXH': '', 'ZDCQTEDGZQY': '', 'FBZDZZXHBJZZ': '', 'CLQXPZYJ': '', 'ZDJJZDXTXH': '',
        'XJXS': '', 'DZZDXTKJQXH': '', 'DWLHSQGRZZ': '', 'DWLHSQZDMHZZ': '', 'HBFHBZ': '', 'GTWBZHXBFHZZJJ': '',
        'CWXLJ': '', 'XLCJG': '', 'XLCYSZZ': '', 'YXCMFH': '', 'YXJQDJL': '', 'JQKYB': '',
        'CDZFHZZ': '', 'FDJWZ': '', 'DCFSLHQ': '', 'TBQZYBZ': '', 'CKMSLJWZ': '', 'CKMJG': '',
        'WTSYJCSL': '', 'YJCXH': '', 'HWYJCKPZ': '', 'YJMSLJWZ': '', 'YJMYDKD': '', 'ZJJ': '',
        'ZYS': '', 'ZDK': '', 'ZYHY': '', 'ZYAQD': '', 'AQDTXZZ': '', 'SPJKXTXH': '',
        'KQTJZZ': '', 'TFHQZZ': '', 'WSJ': '', 'ZDPBQ': '', 'ZDQYXSJXYZZ': '', 'ZKRS': '',
        'XJLX': '', 'CYZWSBS': '', 'RJXLCRJ': '', 'TCXSL': '', 'RYXSL': '', 'YXJHDJL': '',
        'DPJZRHXT': '', 'FDJCZDMHZZ': '', 'DLDCXZDMHZZ': '', 'SDZXJG': '', 'JSYSFDB': '', 'AQBZ': '',
        'CKMYJKZQ': '', 'YJCSL': '', 'YJCSXXHBJZZ': '', 'AQDCSLJWZ': '', 'AQCKSL': '', 'KCNTDK': '',
        'CNTDZDZY': '', 'KBJDKT': '', 'KBG': '', 'FS_KTDC': '', 'ZYJD': '', 'CNXLJ': '',
        'KQJHZZ': '', 'YYBFJMKFSB': '', 'CANZX': '', 'ZCZZXH': '', 'ZDQXS_QH': '', 'JXLJZZXH': '',
        'BGC_QYZGG': '', 'BGCQHZBJ': '', 'BGCJXBJ': '', 'ZDXYSJC': '', 'QYGGHZXKZJ': '', 'GC_QYHCZHDSPJL': '',
        'GCSJZGCS': '', 'CSFGBS': '', 'QZWBJXSJZZ': '', 'WBBZBSL': '', 'CCBZP': '', 'QYXZBLDGD': '',
        'BGC_BGQYCQHZBJ': '', 'BGC_BGQYCHHZBJ': '', 'BGCZHDSPJL': '', 'ZZHGCQHZBJ': '', 'QYGGHZXLDGD': '', 'ZJAQQL': '',
        'RJXCD': '', 'MQJCXT': '', 'DLZX': '', 'ZHJ': '', 'FJAQQL': '', 'TYCXZX': '',
        'DCLD': '', 'MHQ': '', 'ZTGCZZL': '', 'QYZCZMLDGD': '', 'BGQYCQHZBJ': '', 'BGQYCHHZBJ': '',
        'ZDXYSJA': '', 'BGQYCLZQDDJL': '', 'QYGLJQXH': '', 'QYHCZHDSPJL': '', 'WXHWYSCLLX': '', 'HXLBNCC': '',
        'YZWTSYZZ': '', 'BGQYC_BGCQYXGG': '', 'BGQYC_BGCQHZBJ': '', 'BGQYC_BGCJXBJ': '', 'ZDXYSJB': '',
        'BGQYC_BGCZHDSPJL': '', 'QYGLJQZXLDGD': '', 'QYHC_ZZZGCQHZBJ': ''
    }


# 解析乘用车
def parse_cyc_page(soup):
    data = get_data()
    data['CLLB'] = '乘用车'
    tables = soup.select('.info-table.thCenter.w')
    data['DBCXBH'] = tables[0].find('span').text
    trs = tables[1].findAll('tr')

    data['SCQY'] = trs[0].findAll('td')[1].text

    data['CPXH'] = trs[1].findAll('td')[1].text
    data['CPMC'] = trs[1].findAll('td')[3].text

    data['LEVEL'] = trs[2].findAll('td')[1].text
    data['OUT_SIZE'] = trs[2].findAll('td')[3].text

    data['ZZL'] = trs[3].findAll('td')[1].text
    data['ZBZL'] = trs[3].findAll('td')[3].text

    data['FDJXH'] = trs[4].findAll('td')[1].text
    data['LTGG'] = trs[4].findAll('td')[3].text

    data['ZDQXS_QH'] = trs[5].findAll('td')[1].text
    data['ZHJ'] = trs[5].findAll('td')[3].text

    data['ZYS'] = trs[6].findAll('td')[1].text
    data['ZJJ'] = trs[6].findAll('td')[3].text

    data['ZDK'] = trs[7].findAll('td')[1].text
    data['KBG'] = trs[7].findAll('td')[3].text

    data['ZYAQD'] = trs[8].findAll('td')[1].text
    data['BKJDKT'] = trs[8].findAll('td')[3].text

    data['ZJAQQL'] = trs[9].findAll('td')[1].text
    data['FJAQQL'] = trs[9].findAll('td')[3].text

    data['KQTJZZ'] = trs[10].findAll('td')[1].text
    data['WXDW'] = trs[10].findAll('td')[3].text

    data['RJXCD'] = trs[11].findAll('td')[1].text
    data['LTQYJCXTXH'] = trs[11].findAll('td')[3].text

    data['MQJCXT'] = trs[12].findAll('td')[1].text
    data['TYCXZX'] = trs[12].findAll('td')[3].text

    data['DLZX'] = trs[13].findAll('td')[1].text
    data['DCLD'] = trs[13].findAll('td')[3].text

    data['ESCXH'] = trs[14].findAll('td')[1].text
    data['MHQ'] = trs[14].findAll('td')[3].text

    data['LDWSXH'] = trs[15].findAll('td')[1].text
    data['FBZDZZ'] = trs[15].findAll('td')[3].text

    data['ZDJJZDXTXH'] = trs[16].findAll('td')[1].text
    data['ZHRLXHL'] = trs[16].findAll('td')[3].text
    return data


# 解析客车详情页内容,需要解析车辆类别
def parse_kc_page(soup):
    data = get_data()
    data['CLLB'] = '客车'
    tables = soup.select('.info-table.thCenter.w')
    data['DBCXBH'] = tables[0].find('span').text
    trs = tables[1].findAll('tr')

    data['SCQY'] = trs[0].findAll('td')[1].text

    data['CPXH'] = trs[1].findAll('td')[1].text
    data['CPMC'] = trs[1].findAll('td')[3].text

    data['DPXH'] = trs[2].findAll('td')[1].text
    data['FDJXH'] = trs[2].findAll('td')[3].text

    data['OUT_SIZE'] = trs[3].findAll('td')[1].text
    data['ZKRS'] = trs[3].findAll('td')[3].text

    data['KCLX'] = trs[4].findAll('td')[1].text
    data['LEVEL'] = trs[4].findAll('td')[3].text

    data['ZZL'] = trs[5].findAll('td')[1].text
    data['ZBZL'] = trs[5].findAll('td')[3].text

    data['RLZL'] = trs[6].findAll('td')[1].text
    data['ZGCS'] = trs[6].findAll('td')[3].text

    data['LTGG'] = trs[7].findAll('td')[1].text
    data['LTSL'] = trs[7].findAll('td')[3].text

    data['ZDQXS_QH'] = trs[8].findAll('td')[1].text
    data['XJLX'] = trs[8].findAll('td')[3].text

    data['CWXLJ'] = trs[9].findAll('td')[1].text
    data['CYZWSBS'] = trs[9].findAll('td')[3].text

    data['XLCJG'] = trs[10].findAll('td')[1].text
    data['RJXLCRJ'] = trs[10].findAll('td')[3].text

    data['XLCYSZZ'] = trs[11].findAll('td')[1].text
    data['TCXSL'] = trs[11].findAll('td')[3].text

    data['YXCMFH'] = trs[12].findAll('td')[1].text
    data['RYXSL'] = trs[12].findAll('td')[3].text

    data['YXJQDJL'] = trs[13].findAll('td')[1].text
    data['YXJHDJL'] = trs[13].findAll('td')[3].text

    data['JQKYB'] = trs[14].findAll('td')[1].text
    data['RQPSLJWZ'] = trs[14].findAll('td')[3].text

    data['QYZDXT'] = trs[15].findAll('td')[1].text
    data['ZDJXZDTZZZ'] = trs[15].findAll('td')[3].text

    data['CDZFHZZ'] = trs[16].findAll('td')[1].text
    data['DPJZRHXT'] = trs[16].findAll('td')[3].text

    data['FDJWZ'] = trs[17].findAll('td')[1].text
    data['FDJCZDMHZZ'] = trs[17].findAll('td')[3].text

    data['HSQZDZZ'] = trs[18].findAll('td')[1].text
    data['DLDCXZDMHZZ'] = trs[18].findAll('td')[3].text

    data['DWLHSQBJXT'] = trs[19].findAll('td')[1].text
    data['DWLHSQGRZZ'] = trs[19].findAll('td')[3].text

    data['DCFSLHQ'] = trs[20].findAll('td')[1].text
    data['SDZXJG'] = trs[20].findAll('td')[3].text

    data['TBQZYBZ'] = trs[21].findAll('td')[1].text
    data['JSYSFDB'] = trs[21].findAll('td')[3].text

    data['CKMSLJWZ'] = trs[22].findAll('td')[1].text
    data['AQBZ'] = trs[22].findAll('td')[3].text

    data['CKMJG'] = trs[23].findAll('td')[1].text
    data['CKMYJKZQ'] = trs[23].findAll('td')[3].text

    data['WTSYJCSL'] = trs[24].findAll('td')[1].text
    data['YJCSL'] = trs[24].findAll('td')[3].text

    data['YJCXH'] = trs[25].findAll('td')[1].text
    data['YJCSXXHBJZZ'] = trs[25].findAll('td')[3].text

    data['HWYJCKPZ'] = trs[26].findAll('td')[1].text
    data['AQDCSLJWZ'] = trs[26].findAll('td')[3].text

    data['YJMSLJWZ'] = trs[27].findAll('td')[1].text
    data['AQCKSL'] = trs[27].findAll('td')[3].text

    data['YJMYDKD'] = trs[28].findAll('td')[1].text
    data['KCNTDK'] = trs[28].findAll('td')[3].text

    data['ZJJ'] = trs[29].findAll('td')[1].text
    data['CNTDZDZY'] = trs[29].findAll('td')[3].text

    data['ZYS'] = trs[30].findAll('td')[1].text
    data['KBJDKT'] = trs[30].findAll('td')[3].text

    data['ZDK'] = trs[31].findAll('td')[1].text
    data['KBG'] = trs[31].findAll('td')[3].text

    data['ZYHY'] = trs[32].findAll('td')[1].text
    data['FS_KTDC'] = trs[32].findAll('td')[3].text

    data['ZYAQD'] = trs[33].findAll('td')[1].text
    data['ZYJD'] = trs[33].findAll('td')[3].text

    data['AQDTXZZ'] = trs[34].findAll('td')[1].text
    data['CNXLJ'] = trs[34].findAll('td')[3].text

    data['SPJKXTXH'] = trs[35].findAll('td')[1].text
    data['WXDW'] = trs[35].findAll('td')[3].text

    data['KQTJZZ'] = trs[36].findAll('td')[1].text
    data['KQJHZZ'] = trs[36].findAll('td')[3].text

    data['TFHQZZ'] = trs[37].findAll('td')[1].text
    data['YYBFJMKFSB'] = trs[37].findAll('td')[3].text

    data['WSJ'] = trs[38].findAll('td')[1].text
    data['CANZX'] = trs[38].findAll('td')[3].text

    data['ZDPBQ'] = trs[39].findAll('td')[1].text
    data['LGBZXH'] = trs[39].findAll('td')[3].text

    data['ZDQYXSJXYZZ'] = trs[40].findAll('td')[1].text
    data['LTQYJCXTXH'] = trs[40].findAll('td')[3].text

    data['ZDCQTEDGZQY'] = trs[41].findAll('td')[1].text
    data['ZDQCP'] = trs[41].findAll('td')[3].text

    data['ESCXH'] = trs[42].findAll('td')[1].text
    data['FBZDZZXHBJZZ'] = trs[42].findAll('td')[3].text

    data['LDWSXH'] = trs[43].findAll('td')[1].text
    data['FBZDZZ'] = trs[43].findAll('td')[3].text

    data['ZDJJZDXTXH'] = trs[44].findAll('td')[1].text
    data['ZHRLXHL'] = trs[44].findAll('td')[3].text
    return data


# 解析运输车
def parse_ysc_page(soup):
    data = get_data()
    data['CLLB'] = '货车'
    tables = soup.select('.info-table.thCenter.w')
    data['DBCXBH'] = tables[0].find('span').text
    trs = tables[1].findAll('tr')

    data['SCQY'] = trs[0].findAll('td')[1].text

    data['CPXH'] = trs[1].findAll('td')[1].text
    data['CPMC'] = trs[1].findAll('td')[3].text

    data['DPXH'] = trs[2].findAll('td')[1].text
    data['FDJXH'] = trs[2].findAll('td')[3].text

    data['OUT_SIZE'] = trs[3].findAll('td')[1].text
    data['HXLBNCC_RJ'] = trs[3].findAll('td')[3].text

    data['ZZL'] = trs[4].findAll('td')[1].text
    data['ZBZL'] = trs[4].findAll('td')[3].text

    data['RLZL'] = trs[5].findAll('td')[1].text
    data['ZGCS'] = trs[5].findAll('td')[3].text

    data['QDXS'] = trs[6].findAll('td')[1].text
    data['ZXZSL'] = trs[6].findAll('td')[3].text

    data['LTGG'] = trs[7].findAll('td')[1].text
    data['LTSL'] = trs[7].findAll('td')[3].text

    data['CMFHZZ'] = trs[8].findAll('td')[1].text
    data['HXBFH'] = trs[8].findAll('td')[3].text

    data['QXBFHZZ'] = trs[9].findAll('td')[1].text
    data['QZWBJSBZ'] = trs[9].findAll('td')[3].text

    data['JGDSL_QX'] = trs[10].findAll('td')[1].text
    data['QCDJDXJ'] = trs[10].findAll('td')[3].text

    data['JGDSL_SPCZM'] = trs[11].findAll('td')[1].text
    data['RQPSLJWZ'] = trs[11].findAll('td')[3].text

    data['ZHBZBS'] = trs[12].findAll('td')[1].text
    data['QYZDXT'] = trs[12].findAll('td')[3].text

    data['ZDQXS_QH'] = trs[13].findAll('td')[1].text
    data['YLCSLJQ_CQT'] = trs[13].findAll('td')[3].text

    data['ZDJXZDTZZZ'] = trs[14].findAll('td')[1].text
    data['YLCSLJQ_ZDQS'] = trs[14].findAll('td')[3].text

    data['WDJKZZ'] = trs[15].findAll('td')[1].text
    data['WXDW'] = trs[15].findAll('td')[3].text

    data['JSSLTBS'] = trs[16].findAll('td')[1].text
    data['QTXLBJ'] = trs[16].findAll('td')[3].text

    data['LGBZXH'] = trs[17].findAll('td')[1].text
    data['LTQYJCXTXH'] = trs[17].findAll('td')[3].text

    data['ZDQCP'] = trs[18].findAll('td')[1].text
    data['ZDCQTEDGZQY'] = trs[18].findAll('td')[3].text

    data['FBZDZZ'] = trs[19].findAll('td')[1].text
    data['FBZDZZXHBJZZ'] = trs[19].findAll('td')[3].text

    data['ESCXH'] = trs[20].findAll('td')[1].text
    data['CLQXPZYJ'] = trs[20].findAll('td')[3].text

    data['LDWSXH'] = trs[21].findAll('td')[1].text
    data['ZDJJZDXTXH'] = trs[21].findAll('td')[3].text

    data['ZHRLXHL'] = trs[22].findAll('td')[1].text
    return data


# 牵引车页面解析
def parse_qyc_page(soup):
    data = get_data()
    data['CLLB'] = '牵引车辆'
    tables = soup.select('.info-table.thCenter.w')
    data['DBCXBH'] = tables[0].find('span').text
    trs = tables[1].findAll('tr')

    data['SCQY'] = trs[0].findAll('td')[1].text

    data['CPXH'] = trs[1].findAll('td')[1].text
    data['CPMC'] = trs[1].findAll('td')[3].text

    data['DPXH'] = trs[2].findAll('td')[1].text
    data['FDJXH'] = trs[2].findAll('td')[3].text

    data['OUT_SIZE'] = trs[3].findAll('td')[1].text
    data['HXLBNCC'] = trs[3].findAll('td')[3].text

    data['ZZL'] = trs[4].findAll('td')[1].text
    data['ZBZL'] = trs[4].findAll('td')[3].text

    data['ZTGCZZL'] = trs[5].findAll('td')[1].text
    data['ZGCS'] = trs[5].findAll('td')[3].text

    data['RLZL'] = trs[6].findAll('td')[1].text
    data['JXLJZZXH'] = trs[6].findAll('td')[3].text

    data['QDXS'] = trs[7].findAll('td')[1].text
    data['ZXZ'] = trs[7].findAll('td')[3].text

    data['LTGG'] = trs[8].findAll('td')[1].text

    data['LTSL'] = trs[9].findAll('td')[1].text
    data['QXBFHZZ'] = trs[9].findAll('td')[3].text

    data['CMFHZZ'] = trs[10].findAll('td')[1].text
    data['HXBFH'] = trs[10].findAll('td')[3].text

    data['CSFGBS'] = trs[11].findAll('td')[1].text
    data['WBBZBSL'] = trs[11].findAll('td')[3].text

    data['QZWBJSBZ'] = trs[12].findAll('td')[1].text
    data['QZWBJXSJZZ'] = trs[12].findAll('td')[3].text

    data['HSQZDZZ'] = trs[13].findAll('td')[1].text
    data['DWLHSQZDMHZZ'] = trs[13].findAll('td')[3].text

    data['DWLHSQGRZZ'] = trs[14].findAll('td')[1].text
    data['DWLHSQBJXT'] = trs[14].findAll('td')[3].text

    data['JGDSL_QX'] = trs[15].findAll('td')[1].text
    data['QCDJDXJ'] = trs[15].findAll('td')[3].text

    data['JGDSL_SPCZM'] = trs[16].findAll('td')[1].text
    data['RQPSLJWZ'] = trs[16].findAll('td')[3].text

    data['ZHBZBS'] = trs[17].findAll('td')[1].text
    data['QYZDXT'] = trs[17].findAll('td')[3].text

    data['ZDQXS'] = trs[18].findAll('td')[1].text
    data['YLCSLJQ_CQT'] = trs[18].findAll('td')[3].text

    data['ZDJXZDTZZZ'] = trs[19].findAll('td')[1].text
    data['YLCSLJQ_ZDQS'] = trs[19].findAll('td')[3].text

    data['WDJKZZ'] = trs[20].findAll('td')[1].text
    data['WXDW'] = trs[20].findAll('td')[3].text

    data['JSSLTBS'] = trs[21].findAll('td')[1].text
    data['QTXLBJ'] = trs[21].findAll('td')[3].text

    data['LGBZXH'] = trs[22].findAll('td')[1].text
    data['LTQYJCXTXH'] = trs[22].findAll('td')[3].text

    data['ZDQCP'] = trs[23].findAll('td')[1].text
    data['ZDCQTEDGZQY'] = trs[23].findAll('td')[3].text

    data['FBZDZZ'] = trs[24].findAll('td')[1].text
    data['FBZDZZXHBJZZ'] = trs[24].findAll('td')[3].text

    data['ESCXH'] = trs[25].findAll('td')[1].text
    data['CLQXPZYJ'] = trs[25].findAll('td')[3].text

    data['LDWSXH'] = trs[26].findAll('td')[1].text
    data['ZDJJZDXTXH'] = trs[26].findAll('td')[3].text

    data['DZZDXTKJQXH'] = trs[27].findAll('td')[1].text
    data['YZWTSYZZ'] = trs[27].findAll('td')[3].text

    data['QYZCZMLDGD'] = trs[28].findAll('td')[1].text
    data['BGQYC_BGCQYXGG'] = trs[28].findAll('td')[3].text

    data['BGQYCQHZBJ'] = trs[29].findAll('td')[1].text
    data['BGQYC_BGCQHZBJ'] = trs[29].findAll('td')[3].text

    data['BGQYCHHZBJ'] = trs[30].findAll('td')[1].text
    data['BGQYC_BGCJXBJ'] = trs[30].findAll('td')[3].text

    data['ZDXYSJA'] = trs[31].findAll('td')[1].text
    data['ZDXYSJB'] = trs[31].findAll('td')[3].text

    data['BGQYCLZQDDJL'] = trs[32].findAll('td')[1].text
    data['BGQYC_BGCZHDSPJL'] = trs[32].findAll('td')[3].text

    data['QYGLJQXH'] = trs[33].findAll('td')[1].text
    data['QYGLJQZXLDGD'] = trs[33].findAll('td')[3].text

    data['QYHCZHDSPJL'] = trs[34].findAll('td')[1].text
    data['QYHC_ZZZGCQHZBJ'] = trs[34].findAll('td')[3].text

    data['ZHRLXHL'] = trs[35].findAll('td')[1].text

    return data


# 解析挂车页面
def parse_gc_page(soup):
    data = get_data()
    data['CLLB'] = '挂车'
    tables = soup.select('.info-table.thCenter.w')
    data['DBCXBH'] = tables[0].find('span').text
    trs = tables[1].findAll('tr')

    data['SCQY'] = trs[0].findAll('td')[1].text

    data['CPXH'] = trs[1].findAll('td')[1].text
    data['CPMC'] = trs[1].findAll('td')[3].text

    data['OUT_SIZE'] = trs[2].findAll('td')[1].text
    data['HXLBNCC_RJ'] = trs[2].findAll('td')[3].text

    data['ZZL'] = trs[3].findAll('td')[1].text
    data['ZBZL'] = trs[3].findAll('td')[3].text

    data['ZCZZXH'] = trs[4].findAll('td')[1].text
    data['ZXZ'] = trs[4].findAll('td')[3].text

    data['LTGG'] = trs[5].findAll('td')[1].text

    data['LTSL'] = trs[6].findAll('td')[1].text
    data['ZGCS'] = trs[6].findAll('td')[3].text

    data['CMFHZZ'] = trs[7].findAll('td')[1].text
    data['HXBFH'] = trs[7].findAll('td')[3].text

    data['ZHBZBS'] = trs[8].findAll('td')[1].text
    data['CSFGBS'] = trs[8].findAll('td')[1].text

    data['QZWBJSBZ'] = trs[9].findAll('td')[1].text
    data['QZWBJXSJZZ'] = trs[9].findAll('td')[3].text

    data['JGDSL_QX'] = trs[10].findAll('td')[1].text
    data['WBBZBSL'] = trs[10].findAll('td')[3].text

    data['JGDSL_SPCZM'] = trs[11].findAll('td')[1].text
    data['WDJKZZ'] = trs[11].findAll('td')[3].text

    data['ZDQXS_QH'] = trs[12].findAll('td')[1].text
    data['QYZDXT'] = trs[12].findAll('td')[3].text

    data['ZDJXZDTZZZ'] = trs[13].findAll('td')[1].text
    data['YLCSLJQ_CQT'] = trs[13].findAll('td')[3].text

    data['ZDQCP'] = trs[14].findAll('td')[1].text
    data['YLCSLJQ_ZDQS'] = trs[14].findAll('td')[3].text

    data['DZZDXTKJQXH'] = trs[15].findAll('td')[1].text
    data['FBZDZZ'] = trs[15].findAll('td')[3].text

    data['JXLJZZXH'] = trs[16].findAll('td')[1].text
    data['CCBZP'] = trs[16].findAll('td')[3].text

    data['BGC_QYZGG'] = trs[17].findAll('td')[1].text
    data['QYXZBLDGD'] = trs[17].findAll('td')[3].text

    data['BGCQHZBJ'] = trs[18].findAll('td')[1].text
    data['BGC_BGQYCQHZBJ'] = trs[18].findAll('td')[3].text

    data['BGCJXBJ'] = trs[19].findAll('td')[1].text
    data['BGC_BGQYCHHZBJ'] = trs[19].findAll('td')[3].text

    data['ZDXYSJC'] = trs[20].findAll('td')[1].text
    data['BGCZHDSPJL'] = trs[20].findAll('td')[3].text

    data['QYGGHZXKZJ'] = trs[21].findAll('td')[1].text
    data['ZZHGCQHZBJ'] = trs[21].findAll('td')[3].text

    data['GC_QYHCZHDSPJL'] = trs[22].findAll('td')[1].text
    data['QYGGHZXLDGD'] = trs[22].findAll('td')[3].text

    return data


# 爬取数据
def spider():
    link_list = []
    wb = openpyxl.load_workbook(r'C:\Users\13099\Desktop\达标公示数据.xlsx')
    sheet = wb.active
    # 将连接保存到集合
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1).value
        link_list.append(cell)

    # 定义空的结果集合
    result = []
    # 获取到结果集
    index = 0
    for link in link_list:
        print('第', index, '条', '共', len(link_list), '条：', link)
        # 获取到每一页的连接
        response = requests.get(link).content.decode('utf-8')
        soup = BeautifulSoup(response, 'html.parser')
        # 获取标题
        title = soup.find('h3').text
        res = None
        if '客车' in title:
            res = parse_kc_page(soup)
        elif '乘用' in title:
            res = parse_cyc_page(soup)
        elif '牵引' in title:
            res = parse_qyc_page(soup)
        elif '载货汽车' in title:
            res = parse_ysc_page(soup)
        elif '挂车' in title:
            res = parse_gc_page(soup)


        # 添加批次信息
        res['PC'] = 38


        result.append(res)
        index = index + 1

    # 将数据存入excel中
    utils.saveData(result, 'F://达标公示数据.xls')


if __name__ == '__main__':
    spider()
